from docx import Document
import pandas as pd
import os
from openpyxl import load_workbook

def leer_archivo_word(ruta):
    try:
        # Cargar el documento Word
        doc = Document(ruta)
        contenido = []

        # Leer las tablas del documento
        for tabla in doc.tables:
            for fila in tabla.rows:
                fila_texto = [celda.text.strip() for celda in fila.cells if celda.text.strip()]
                if fila_texto:  # Verificar que la fila no esté vacía
                    contenido.append(fila_texto)  # Guardar cada fila como una lista

        return contenido  # Devolver el contenido como un array

    except Exception as e:
        print(f"Ocurrió un error: {e}")
        return []

def guardar_en_excel(data, nombre_archivo):
    # Crear un DataFrame a partir de los datos
    df = pd.DataFrame(data, columns=["N°", "Título", "Contexto", "Resultado / Comportamiento esperado", "Detalle Campos"])

    # Guardar el DataFrame en un archivo Excel
    df.to_excel(nombre_archivo, index=False)

def copiar_datos_entre_archivos(archivo_origen, archivo_destino):
    try:
        # Cargar el archivo origen (resultado.xlsx)
        wb_origen = load_workbook(archivo_origen)
        hoja_origen = wb_origen.active

        # Leer las celdas B4, C4, D4 y E4
        b4 = hoja_origen['B4'].value
        c4 = hoja_origen['C4'].value
        d4 = hoja_origen['D4'].value
        e4 = hoja_origen['E4'].value

        # Cargar el archivo destino (test.xlsx)
        wb_destino = load_workbook(archivo_destino)
        hoja_destino = wb_destino.active

        # Escribir los valores en las celdas B7, C7, D7 y E7
        hoja_destino['B7'] = b4
        hoja_destino['C7'] = c4
        hoja_destino['D7'] = d4
        hoja_destino['E7'] = e4

        # Guardar el archivo destino con los nuevos valores
        wb_destino.save(archivo_destino)

        print(f"Datos copiados de {archivo_origen} a {archivo_destino}")

    except Exception as e:
        print(f"Ocurrió un error al copiar datos: {e}")

if __name__ == "__main__":
    # Ruta del archivo Word
    ruta_archivo = r"C:\Users\d4vid\OneDrive\Escritorio\Generate CP-IA\Archivos_imp\test.docx"
    contenido = leer_archivo_word(ruta_archivo)

    # Obtener la carpeta del archivo Word
    carpeta = os.path.dirname(ruta_archivo)
    ruta_excel_resultado = os.path.join(carpeta, "resultado.xlsx")
    ruta_excel_test = os.path.join(carpeta, "test.xlsx")

    # Guardar el contenido en Excel (resultado.xlsx)
    guardar_en_excel(contenido, ruta_excel_resultado)

    # Copiar datos de resultado.xlsx a test.xlsx
    copiar_datos_entre_archivos(ruta_excel_resultado, ruta_excel_test)

    # Mostrar el contenido por consola
    for linea in contenido:
        print(linea)
