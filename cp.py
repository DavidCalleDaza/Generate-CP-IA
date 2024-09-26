import openpyxl
import shutil
import os
from openpyxl.styles import PatternFill

# Rutas de origen y directorio de destino
source_file = r'C:\Users\d4vid\OneDrive\Escritorio\Generate-CP-IA\Archivos_imp\test.xlsx'
destination_directory = r'C:\Users\d4vid\OneDrive\Escritorio\Generate-CP-IA\Archivos_xpo'

# Obtener el nombre base del archivo original sin extensión
filename = os.path.splitext(os.path.basename(source_file))[0]

# Generar el nuevo nombre con el sufijo 'V1' y la extensión '.xlsx'
new_filename = f"{filename}_V1.xlsx"
destination_file = os.path.join(destination_directory, new_filename)

# Copiar archivo de origen a destino con el nuevo nombre
shutil.copyfile(source_file, destination_file)

# Abrir el archivo copiado para trabajar en él
wb = openpyxl.load_workbook(destination_file)
sheet_historia = wb['HISTORIA DE USUARIO']
sheet_casos = wb['CASOS DE PRUEBA']

# Estilos de color para las pruebas
positive_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde
negative_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")    # Naranja claro

# Número de criterio y ÉPICA, ID_HU desde HISTORIA DE USUARIO
epica = sheet_historia['C1'].value
id_hu = sheet_historia['C2'].value
descripcion_hu = sheet_historia['C3'].value
roles_hu = sheet_historia['C4'].value
quiero_hu = sheet_historia['C5'].value
para_hu = sheet_historia['E1'].value
prerequisitos_hu = sheet_historia['E2'].value

# Función para generar pasos de ejecución basados en CONTEXTO
def generar_pasos(contexto):
    pasos = contexto.split('>')
    pasos_ejecucion = "Pasos específicos para recrear el escenario:\n"
    for i, paso in enumerate(pasos, start=1):
        pasos_ejecucion += f"{i}. {paso.strip()}\n"
    return pasos_ejecucion.strip()

# Función para rellenar los casos de prueba
def generar_cp(sheet, row, tipo_prueba, titulo, contexto, resultado_esperado, detalle_campos, codigo_cp, color, rol):
    # Aquí se asegura que el título se use correctamente
    sheet[f'C{row}'] = titulo  # Título del caso de prueba
    sheet[f'D{row}'] = f"{descripcion_hu} - Rol: {rol}\nQuiero: {quiero_hu}\nPara: {para_hu}\nPrerrequisitos: {prerequisitos_hu}"
    sheet[f'E{row}'] = generar_pasos(contexto)  # Generar pasos basados en CONTEXTO
    sheet[f'F{row}'] = resultado_esperado
    sheet[f'G{row}'] = tipo_prueba
    sheet[f'G{row}'].fill = color  # Aplicar el color de la celda
    sheet[f'H{row}'] = codigo_cp
    sheet[f'I{row}'] = "Pendiente"

# Iterar sobre los criterios en la hoja HISTORIA DE USUARIO (a partir de la fila 7)
row_historia = 7
row_casos = 32  # Fila inicial para los casos de prueba

while sheet_historia[f'B{row_historia}'].value:
    titulo = sheet_historia[f'B{row_historia}'].value  # Asegúrate de que esto sea correcto
    contexto = sheet_historia[f'C{row_historia}'].value
    resultado_esperado = sheet_historia[f'D{row_historia}'].value
    detalle_campos = sheet_historia[f'F{row_historia}'].value
    
    # Crear el código CP: ÉPICA + ID_HU + Número de criterio
    codigo_cp_base = f"{epica}-{id_hu}-{row_historia-6}"
    
    # Separar los roles
    roles = roles_hu.split('-')
    
    # Generar diferentes tipos de casos de prueba (Positivo, Negativo, Límite, Extremo) por cada rol
    for rol in roles:
        rol = rol.strip()  # Eliminar espacios en blanco alrededor del rol
        
        # CP Positivo
        generar_cp(sheet_casos, row_casos, "Positiva", titulo, contexto, resultado_esperado, detalle_campos,
                   codigo_cp_base + "-P", positive_fill, rol)
        row_casos += 1
        
        # CP Negativo
        generar_cp(sheet_casos, row_casos, "Negativa", titulo, contexto, f"Validación negativa de {resultado_esperado}", detalle_campos,
                   codigo_cp_base + "-N", negative_fill, rol)
        row_casos += 1
        
        # CP Límite
        generar_cp(sheet_casos, row_casos, "Límites", titulo, contexto, f"Validación de límite para {resultado_esperado}", detalle_campos,
                   codigo_cp_base + "-L", orange_fill, rol)
        row_casos += 1
        
        # CP Extremo
        generar_cp(sheet_casos, row_casos, "Extremos", titulo, contexto, f"Validación extrema para {resultado_esperado}", detalle_campos,
                   codigo_cp_base + "-E", orange_fill, rol)
        row_casos += 1
    
    # Pasar a la siguiente fila en HISTORIA DE USUARIO
    row_historia += 1

# Guardar los cambios en el archivo
wb.save(destination_file)
print(f"Casos de prueba generados correctamente en {new_filename}.")
