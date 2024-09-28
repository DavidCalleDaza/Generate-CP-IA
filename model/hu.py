from docx import Document
import pandas as pd
import os
from openpyxl import load_workbook

def leer_archivo_word(ruta):
    try:
        # Cargar el documento Word
        doc = Document(ruta)
        contenido = []

        # Leer las tablas del documento
        for tabla in doc.tables:
            for fila in tabla.rows:
                fila_texto = [celda.text.strip() for celda in fila.cells if celda.text.strip()]
                # Asegurarse de que la fila tiene el número esperado de columnas (5 en este caso)
                if fila_texto and len(fila_texto) >= 4:  # Verifica que la fila no esté vacía y tenga al menos 4 elementos
                    contenido.append(fila_texto)  # Guardar cada fila como una lista

        return contenido  # Devolver el contenido como un array

    except Exception as e:
        print(f"Ocurrió un error: {e}")
        return []

def guardar_en_excel(data, nombre_archivo):
    # Filtrar los datos para conservar solo aquellos donde la primera columna (N°) es numérica
    data_filtrada = [fila for fila in data if fila and fila[0].isdigit()]

    # Crear un DataFrame a partir de los datos filtrados
    df = pd.DataFrame(data_filtrada, columns=["N°", "Título", "Contexto", "Resultado / Comportamiento esperado", "Detalle Campos"])

    # Guardar el DataFrame en un archivo Excel
    df.to_excel(nombre_archivo, index=False)

def copiar_criterios_al_test(archivo_origen, archivo_destino):
    try:
        # Cargar el archivo origen (resultado.xlsx) donde se guardaron los criterios
        wb_origen = load_workbook(archivo_origen)
        hoja_origen = wb_origen.active

        # Cargar el archivo destino (test.xlsx)
        wb_destino = load_workbook(archivo_destino)

        # Seleccionar la hoja "HISTORIA DE USUARIO" en el archivo destino
        if "HISTORIA DE USUARIO" in wb_destino.sheetnames:
            hoja_destino = wb_destino["HISTORIA DE USUARIO"]
        else:
            print("La hoja 'HISTORIA DE USUARIO' no existe en el archivo destino.")
            return

        # Leer las celdas del archivo origen y escribirlas a partir de la fila 7 en el archivo destino
        for i, row in enumerate(hoja_origen.iter_rows(min_row=2, max_row=hoja_origen.max_row, values_only=True), start=7):
            # Verificar que la fila no esté vacía y que contenga los datos necesarios
            if row[1] is not None:  # Título no puede ser None
                hoja_destino[f'B{i}'] = row[1]  # Título (columna B)
                hoja_destino[f'C{i}'] = row[2]  # Contexto (columna C)
                hoja_destino[f'D{i}'] = row[3]  # Resultado / Comportamiento esperado (columna D)
                hoja_destino[f'E{i}'] = row[4]  # Detalle Campos (columna E)

        # Guardar el archivo destino con los criterios copiados
        wb_destino.save(archivo_destino)
        print(f"Criterios copiados correctamente en la hoja 'HISTORIA DE USUARIO' de {archivo_destino}")

    except Exception as e:
        print(f"Ocurrió un error al copiar los criterios: {e}")

if __name__ == "__main__":
    # Ruta del archivo Word
    ruta_archivo = r"C:\Users\ASUS\Documents\GitHub\Generate-CP-IA\Archivos_imp\Criterios_de_aceptacion.docx"
    contenido = leer_archivo_word(ruta_archivo)

    # Obtener la carpeta del archivo Word
    carpeta = os.path.dirname(ruta_archivo)
    ruta_excel_resultado = os.path.join(carpeta, "resultado.xlsx")
    ruta_excel_test = os.path.join(carpeta, "test.xlsx")

    # Guardar el contenido en Excel (resultado.xlsx) filtrando registros no numéricos
    guardar_en_excel(contenido, ruta_excel_resultado)

    # Comentar o eliminar la llamada a la función para copiar criterios
    # copiar_criterios_al_test(ruta_excel_resultado, ruta_excel_test)

    # Mostrar el contenido por consola
    for linea in contenido:
        print(linea)
