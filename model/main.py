import subprocess
import os
from openpyxl import load_workbook  # Asegúrate de tener openpyxl instalado: pip install openpyxl

# Ruta del archivo test.xlsx
test_file_path = r"C:\Users\ASUS\Documents\GitHub\Generate-CP-IA\Archivos_imp\test.xlsx"

def limpiar_test_file():
    """Limpia el archivo test.xlsx eliminando las filas desde la fila 7 en adelante."""
    try:
        # Cargar el libro de Excel
        wb = load_workbook(test_file_path)
        # Seleccionar la hoja 'HISTORIA DE USUARIO'
        if 'HISTORIA DE USUARIO' in wb.sheetnames:
            sheet = wb['HISTORIA DE USUARIO']
            # Eliminar filas desde la 7 en adelante
            sheet.delete_rows(7, sheet.max_row - 6)  # Desde la fila 7 hasta la última fila
            # Guardar los cambios
            wb.save(test_file_path)
            print("Información borrada de test.xlsx desde la fila 7 en adelante.")
        else:
            print("La hoja 'HISTORIA DE USUARIO' no se encuentra en el archivo test.xlsx.")
    except Exception as e:
        print(f"Error al limpiar test.xlsx: {e}")

def main():
    try:
        # 1. Ejecutar la función limpiar_test_file()
        limpiar_test_file()

        # 2. Rutas completas a los scripts
        cargar_hu_path = r"C:\Users\ASUS\Documents\GitHub\Generate-CP-IA\model\cargar_hu.py"
        hu_path = r"C:\Users\ASUS\Documents\GitHub\Generate-CP-IA\model\hu.py"

        # 3. Ejecutar el script cargar_hu.py
        subprocess.run(['python', cargar_hu_path], check=True)
        print("Ejecutado: cargar_hu.py")

        # 4. Ejecutar el script hu.py
        subprocess.run(['python', hu_path], check=True)
        print("Ejecutado: hu.py")

        # 5. Abrir el archivo CP-VBA.xlsm utilizando Excel
        file_path = r"C:\Users\ASUS\Documents\GitHub\Generate-CP-IA\Archivos_xpo\VBA\CP-VBA.xlsm"
        subprocess.run(['start', 'excel', file_path], shell=True)  # Abrir usando Excel
        print(f"Abrir archivo: {file_path}")

    except subprocess.CalledProcessError as e:
        print(f"Error al ejecutar el script: {e}")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

if __name__ == "__main__":
    main()
