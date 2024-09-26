import openpyxl

# Rutas de los archivos
source_file = r"C:\Users\d4vid\OneDrive\Escritorio\Generate-CP-IA\Archivos_imp\test.xlsx"
destination_file = r"C:\Users\d4vid\OneDrive\Escritorio\Generate-CP-IA\Archivos_xpo\CP-VBA.xlsm"

# Cargar el archivo de origen y destino
source_wb = openpyxl.load_workbook(source_file)
destination_wb = openpyxl.load_workbook(destination_file)

# Seleccionar las hojas de interés
source_sheet = source_wb["HISTORIA DE USUARIO"]
destination_sheet = destination_wb["HISTORIA DE USUARIO"]

# Limpiar los datos previos en las celdas B7:E del destino
for row in destination_sheet.iter_rows(min_row=7, max_row=destination_sheet.max_row, min_col=2, max_col=5):
    for cell in row:
        cell.value = None

# Copiar registros desde la fila 7 en adelante
for idx, row in enumerate(source_sheet.iter_rows(min_row=7, max_row=source_sheet.max_row, max_col=5, values_only=True), start=7):
    destination_sheet.cell(row=idx, column=2, value=row[0])  # B
    destination_sheet.cell(row=idx, column=3, value=row[1])  # C
    destination_sheet.cell(row=idx, column=4, value=row[2])  # D
    destination_sheet.cell(row=idx, column=5, value=row[3])  # E

# Guardar cambios en el archivo de destino
destination_wb.save(destination_file)

# Cerrar los archivos
source_wb.close()
destination_wb.close()

print("Registros copiados exitosamente.")
